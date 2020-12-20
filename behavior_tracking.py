import matplotlib.pyplot as plt
import pandas as pd
import xlwt
from xlwt import Workbook
import xlsxwriter
import time as time

print("\t\tUser Behavior Tracking Application\n")
time.sleep(2)


def application(userAction):
    if userAction.lower() == 'record':
        print('studying,sleeping,eating,leisure,exercise,socialMedia ')
        # total=int(input('for whow many activities you want to keep track maxvalue =6:'))
        list1 = ['start', 'date']
        activity_list = ['start', 'time', 'study', 'sleeping', 'eating', 'leisure', 'exercise', 'socialMedia']
        activity = ''
        print('study,sleeping,eating,leisure,exercise,socialMedia ')
        # entering exact activity name mentioned in the list
        '''
        for i in range(total):
            activity=input('pls enter exact activity name : ')
            if activity not in activity_list:
                break
            else:
                list1.append(activity)
        print(list1)
        '''
        from openpyxl import load_workbook

        filename = 'Activity_Tracker.xlsx'

        wb = load_workbook(filename)

        ws = wb.worksheets[0]
        for i in range(1, len(activity_list)):
            # ws.write(0,i,list1[i])
            ws.cell(row=1, column=i).value = activity_list[i]
            ws['K1'] = len(activity_list)

        wb.save(filename)

        # ===================================================================
        # ENTERING DATA
        time_calculator = [0]
        total_time = 0
        time = 0
        count = 2
        # k=ws['K1'].value
        # print(k)
        print('enter "date,study,sleep,eating,leisure,exercise,socialMedia" in values in sequence')
        for j in range(7):
            # print(list1[j])
            if j == 0:
                time = ''
                date = input('Enter date in dd format:')
                month = input('Enter month in mm format:')
                year = input('Enter year in yyyy:')
                time = date + '-' + month + '-' + year
                time_calculator.append(time)
            else:
                time = int(input('Enter time in minutes :'))
                print()
                time_calculator.append(time)
                total_time += time

        print(time_calculator)
        print('Total time spent in all activities :', total_time)
        # ENTERING USERDATA IN EXCEL

        filename = 'Activity_Tracker.xlsx'

        wb = load_workbook(filename)

        ws = wb.worksheets[0]
        for i in range(1, len(time_calculator)):
            while ws.cell(row=count, column=i).value != None:
                count = count + 1
            # ws.write(0,i,list1[i])
            ws.cell(row=count, column=i).value = time_calculator[i]

        wb.save(filename)
        # Workbook is created
    elif userAction.lower() == "visualize":
        df = pd.read_excel(r"C:\Users\Siva Pavan\PycharmProjects\AML_1204\Activity_Tracker.xlsx")

        join = input('Do you want a Graph Representation(yes/no):')
        if join.lower() == 'yes':
            print("\tFull Activity Report (FAR) or Activity Wise Report (WAR)")
            join1 = input("Which report you want FAR/WAR:")
            if join1.lower() == "far":
                fig, axes = plt.subplots(nrows=2, ncols=3, figsize=(15, 15))
                plt.subplots_adjust(wspace=0.30, hspace=0.30, top=0.92)
                plt.suptitle("Full Activity Report", fontsize=15)
                axes[0, 0].hist(df.study)
                axes[0, 0].set_xlabel("time")
                axes[0, 0].set_ylabel("study")
                axes[0, 1].hist(df.sleeping)
                axes[0, 1].set_xlabel("time")
                axes[0, 1].set_ylabel("sleeping")
                axes[0, 2].hist(df.eating)
                axes[0, 2].set_xlabel("time")
                axes[0, 2].set_ylabel("eating")
                axes[1, 0].hist(df.leisure)
                axes[1, 0].set_xlabel("time")
                axes[1, 0].set_ylabel("leisure")
                axes[1, 1].hist(df.exercise)
                axes[1, 1].set_xlabel("time")
                axes[1, 1].set_ylabel("exercise")
                axes[1, 2].hist(df.socialMedia)
                axes[1, 2].set_xlabel("time")
                axes[1, 2].set_ylabel("socialMedia")
                plt.show()

            elif join1.lower() == "war":
                join2 = input("which graph(pie/bar):")

                if join2.lower() == 'pie':
                    print("pie chart")

                    def pieGraph(activityName):
                        if activityName == 'study':
                            labels = df.study.unique()
                            colors = ['olivedrab', 'yellowgreen']
                            values = df.study.value_counts().values
                            plt.figure(figsize=(7, 7))
                            plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%')
                            plt.title('Monthly Analysis', color='black', fontsize=10)
                            plt.show()
                        elif activityName == 'sleeping':
                            labels = df.sleeping.unique()
                            colors = ['olivedrab', 'yellowgreen']
                            values = df.sleeping.value_counts().values
                            plt.figure(figsize=(7, 7))
                            plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%')
                            plt.title('Monthly Analysis', color='black', fontsize=10)
                            plt.show()
                        elif activityName == 'eating':
                            labels = df.eating.unique()
                            colors = ['olivedrab', 'yellowgreen']
                            values = df.eating.value_counts().values
                            plt.figure(figsize=(7, 7))
                            plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%')
                            plt.title('Monthly Analysis', color='black', fontsize=10)
                            plt.show()
                        elif activityName == 'leisure':
                            labels = df.leisure.unique()
                            colors = ['olivedrab', 'yellowgreen']
                            values = df.leisure.value_counts().values
                            plt.figure(figsize=(7, 7))
                            plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%')
                            plt.title('Monthly Analysis', color='black', fontsize=10)
                            plt.show()
                        elif activityName == 'exercise':
                            labels = df.exercise.unique()
                            colors = ['olivedrab', 'yellowgreen']
                            values = df.exercise.value_counts().values
                            plt.figure(figsize=(7, 7))
                            plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%')
                            plt.title('Monthly Analysis', color='black', fontsize=10)
                            plt.show()
                        elif activityName == 'socialMedia':
                            labels = df.socialMedia.unique()
                            colors = ['olivedrab', 'yellowgreen']
                            values = df.socialMedia.value_counts().values
                            plt.figure(figsize=(7, 7))
                            plt.pie(values, labels=labels, colors=colors, autopct='%1.1f%%')
                            plt.title('Monthly Analysis', color='black', fontsize=10)
                            plt.show()
                        else:
                            print("Wrong Input for Activity")

                    k1 = input("Enter Activity Name:")
                    pieGraph(k1)

                elif join2.lower() == 'bar':
                    print("bar chart")

                    def barGraph(activity1):
                        if activity1 == 'study':
                            plt.figure(1, figsize=(8, 5))
                            plt.bar(df.study.unique(), height=df.study.value_counts(), color='red')
                            plt.xlabel("Unique Values")
                            plt.show()
                        elif activity1 == 'sleeping':
                            plt.figure(1, figsize=(8, 5))
                            plt.bar(df.sleeping.unique(), height=df.sleeping.value_counts(), color='red')
                            plt.xlabel("Unique Values")
                            plt.show()
                        elif activity1 == 'eating':
                            plt.figure(1, figsize=(8, 5))
                            plt.bar(df.eating.unique(), height=df.eating.value_counts(), color='red')
                            plt.xlabel("Unique Values")
                            plt.show()
                        elif activity1 == 'leisure':
                            plt.figure(1, figsize=(8, 5))
                            plt.bar(df.leisure.unique(), height=df.leisure.value_counts(), color='red')
                            plt.xlabel("Unique Values")
                            plt.show()
                        elif activity1 == 'exercise':
                            plt.figure(1, figsize=(8, 5))
                            plt.bar(df.exercise.unique(), height=df.exercise.value_counts(), color='red')
                            plt.xlabel("Unique Values")
                            plt.show()
                        elif activity1 == 'socialMedia':
                            plt.figure(1, figsize=(8, 5))
                            plt.bar(df.socialMedia.unique(), height=df.socialMedia.value_counts(), color='red')
                            plt.xlabel("Unique Values")
                            plt.show()
                        else:
                            print("Wrong Input for Activity")

                    k2 = input("Enter Activity Name:")
                    barGraph(k2)

                else:
                    print("Wrong Input")
            else:
                print("Wrong Input")
        elif join == 'no':
            print("Thank You For Using Application")
        else:
            print("Wrong Input")
    else:
        print("Wrong Input")


print("\tDo you want to record or visualize data")
k = input("Record/Visualize:")
application(k)
time.sleep(1)
print("Thank You For Using Application")
